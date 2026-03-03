#!/usr/bin/env python3
"""Generate SPAWT individual form sheets — one A4 page per lieu."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from datetime import date

wb = Workbook()

# ── Palette ──
GOLD = "B8960C"
DARK = "1A1A1A"
WHITE = "FFFFFF"
BG = "F8F7F4"
GREEN = "5B8C5A"
MUTED = "8B8B8B"
LIGHT_GOLD = "FFF8E7"
LIGHT_GREEN = "EFF6EE"
LIGHT_BLUE = "EBF0FA"
LIGHT_PINK = "FDF0F0"
LIGHT_PURPLE = "F3EEFA"
LIGHT_GREY = "F5F5F5"

# ── Styles ──
title_font = Font(name="Calibri", bold=True, size=18, color=DARK)
section_font = Font(name="Calibri", bold=True, size=11, color=WHITE)
label_font = Font(name="Calibri", bold=True, size=9, color=MUTED)
value_font = Font(name="Calibri", size=11, color=DARK)
hint_font = Font(name="Calibri", size=8, color=MUTED, italic=True)
checkbox_font = Font(name="Calibri", size=9, color=DARK)
small_bold = Font(name="Calibri", bold=True, size=10, color=DARK)

fill_gold = PatternFill("solid", fgColor=GOLD)
fill_dark = PatternFill("solid", fgColor=DARK)
fill_green = PatternFill("solid", fgColor=GREEN)
fill_blue = PatternFill("solid", fgColor="3B82F6")
fill_red = PatternFill("solid", fgColor="DC2626")
fill_purple = PatternFill("solid", fgColor="7C3AED")
fill_grey = PatternFill("solid", fgColor="6B7280")
fill_bg = PatternFill("solid", fgColor=BG)
fill_white = PatternFill("solid", fgColor=WHITE)
fill_light_gold = PatternFill("solid", fgColor=LIGHT_GOLD)
fill_light_green = PatternFill("solid", fgColor=LIGHT_GREEN)
fill_light_blue = PatternFill("solid", fgColor=LIGHT_BLUE)
fill_light_pink = PatternFill("solid", fgColor=LIGHT_PINK)
fill_light_purple = PatternFill("solid", fgColor=LIGHT_PURPLE)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
left_center = Alignment(horizontal="left", vertical="center", wrap_text=True)
right_center = Alignment(horizontal="right", vertical="center")

border_thin = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
border_bottom = Border(bottom=Side(style="thin", color="CCCCCC"))
border_box = Border(
    left=Side(style="medium", color="BBBBBB"),
    right=Side(style="medium", color="BBBBBB"),
    top=Side(style="medium", color="BBBBBB"),
    bottom=Side(style="medium", color="BBBBBB"),
)

# ── Validation data ──
VALIDATIONS = {
    "quartier": "Cocody,Plateau,Marcory,Treichville,Yopougon,Abobo,Adjamé,Koumassi,Port-Bouët,Bingerville,Riviera,2 Plateaux,Angré,Zone 4,Vallon,Bassam,Assinie,Bouaké",
    "type": "Maquis,Restaurant,Gargote,Street food,Bistrot,Fine dining,Snack",
    "cuisine": "Ivoirienne,Africaine (autre),Française,Libanaise,Asiatique,Fast food,Fusion",
    "parking": "Facile,Modéré,Difficile,Impossible",
    "potentiel": "PÉPITE,COUP DE CŒUR,SOLIDE,MOYEN,SKIP",
    "reco": "Absolument,Oui,Peut-être,Non",
    "affluence": "Vide,Calme,Modérée,Bondé,File d'attente",
    "sonore": "Calme,Ambiance musicale,Animé,Bruyant",
    "proprete": "Impeccable,Correcte,Passable,Problématique",
    "accueil": "Chaleureux,Pro,Distant,Désagréable",
    "receptivite": "Très intéressé,Ouvert,Sceptique,Refus",
    "fonction": "Propriétaire,Gérant,Responsable",
}


def add_validation(ws, cell_range, key):
    """Add dropdown validation to a cell range."""
    if key in VALIDATIONS:
        dv = DataValidation(
            type="list",
            formula1=f'"{VALIDATIONS[key]}"',
            allow_blank=True,
            showDropDown=False,
        )
        dv.showInputMessage = True
        dv.showErrorMessage = False
        ws.add_data_validation(dv)
        dv.add(cell_range)


def make_section_header(ws, row, col_start, col_end, title, fill):
    """Create a colored section header spanning columns."""
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=title)
    c.font = section_font
    c.fill = fill
    c.alignment = center
    for ci in range(col_start, col_end + 1):
        ws.cell(row=row, column=ci).fill = fill
        ws.cell(row=row, column=ci).border = border_thin


def make_label_value(ws, row, label_col, label_text, value_col_start, value_col_end,
                     hint="", row_height=22, validation_key=None):
    """Create a label + empty value cell for data entry."""
    # Label
    c = ws.cell(row=row, column=label_col, value=label_text)
    c.font = label_font
    c.alignment = right_center
    c.fill = fill_bg

    # Value cells merged
    if value_col_start != value_col_end:
        ws.merge_cells(start_row=row, start_column=value_col_start, end_row=row, end_column=value_col_end)
    vc = ws.cell(row=row, column=value_col_start)
    vc.font = value_font
    vc.alignment = left_center
    vc.fill = fill_white
    for ci in range(value_col_start, value_col_end + 1):
        ws.cell(row=row, column=ci).border = border_thin

    if hint:
        vc.value = None
        # Use comment-style hint via number format or just leave placeholder
        vc.font = hint_font

    if validation_key:
        col_letter = get_column_letter(value_col_start)
        add_validation(ws, f"{col_letter}{row}", validation_key)

    ws.row_dimensions[row].height = row_height


def make_checkbox_row(ws, row, col_start, col_end, options, label=""):
    """Create a row of checkboxes (□ Option1  □ Option2 ...)."""
    if label:
        c = ws.cell(row=row, column=col_start, value=label)
        c.font = label_font
        c.alignment = right_center
        c.fill = fill_bg
        col_start += 1

    text = "    ".join([f"☐ {opt}" for opt in options])
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=text)
    c.font = checkbox_font
    c.alignment = left_center
    c.fill = fill_white
    for ci in range(col_start, col_end + 1):
        ws.cell(row=row, column=ci).border = border_thin
    ws.row_dimensions[row].height = 22


def make_scale_row(ws, row, label_col, label, value_col, scale_start, scale_end,
                   left_label, right_label, col_end):
    """Create a scale row: Label [__/12] low ←──────→ high."""
    c = ws.cell(row=row, column=label_col, value=label)
    c.font = label_font
    c.alignment = right_center
    c.fill = fill_bg

    vc = ws.cell(row=row, column=value_col)
    vc.font = Font(name="Calibri", bold=True, size=14, color=GOLD)
    vc.alignment = center
    vc.fill = fill_white
    vc.border = border_thin

    # Scale label
    ws.merge_cells(start_row=row, start_column=value_col + 1, end_row=row, end_column=col_end)
    sc = ws.cell(row=row, column=value_col + 1,
                 value=f"  {left_label}  1 ── 2 ── 3 ── 4 ── 5 ── 6 ── 7 ── 8 ── 9 ── 10 ── 11 ── 12  {right_label}")
    sc.font = Font(name="Calibri", size=8, color=MUTED)
    sc.alignment = center
    ws.row_dimensions[row].height = 24


def build_fiche(ws, fiche_num):
    """Build one complete SPAWT fiche on the given worksheet."""
    # ── Page setup ──
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.3, bottom=0.3, header=0.2, footer=0.2)
    ws.print_options.horizontalCentered = True

    # ── Column widths (8 columns, A-H) ──
    widths = [12, 14, 10, 14, 10, 14, 10, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1  # current row

    # ═══════════════════════════════════════════════════
    # HEADER
    # ═══════════════════════════════════════════════════
    ws.merge_cells(f"A{r}:C{r}")
    c = ws.cell(row=r, column=1, value="SPAWT")
    c.font = Font(name="Courier New", bold=True, size=22, color=DARK)
    c.alignment = left_center

    ws.merge_cells(f"D{r}:F{r}")
    c2 = ws.cell(row=r, column=4, value="Fiche Collecte Terrain")
    c2.font = Font(name="Calibri", size=12, color=MUTED)
    c2.alignment = center

    ws.merge_cells(f"G{r}:H{r}")
    c3 = ws.cell(row=r, column=7, value=f"N° {fiche_num:03d}")
    c3.font = Font(name="Calibri", bold=True, size=14, color=GOLD)
    c3.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[r].height = 32
    r += 1

    # Separator
    for ci in range(1, 9):
        ws.cell(row=r, column=ci).border = Border(bottom=Side(style="medium", color=GOLD))
    ws.row_dimensions[r].height = 4
    r += 1

    # ═══════════════════════════════════════════════════
    # BLOC 1 — IDENTIFICATION
    # ═══════════════════════════════════════════════════
    make_section_header(ws, r, 1, 8, "📍  IDENTIFICATION", fill_dark)
    r += 1

    make_label_value(ws, r, 1, "Nom du lieu", 2, 5)
    make_label_value(ws, r, 6, "Date", 7, 8)
    r += 1

    make_label_value(ws, r, 1, "Adresse", 2, 5)
    make_label_value(ws, r, 6, "Spawter", 7, 8)
    r += 1

    make_label_value(ws, r, 1, "Quartier", 2, 3, validation_key="quartier")
    make_label_value(ws, r, 4, "Téléphone", 5, 6)
    make_label_value(ws, r, 7, "GPS", 8, 8)
    r += 1

    make_checkbox_row(ws, r, 1, 8,
                      ["WhatsApp", "Facebook", "Instagram", "Site web"],
                      label="Digital")
    r += 1

    # ═══════════════════════════════════════════════════
    # BLOC 2 — OFFRE & CUISINE
    # ═══════════════════════════════════════════════════
    make_section_header(ws, r, 1, 8, "🍽️  OFFRE & CUISINE", fill_green)
    r += 1

    make_checkbox_row(ws, r, 1, 8,
                      ["Maquis", "Restaurant", "Gargote", "Street food", "Bistrot", "Fine dining", "Snack"],
                      label="Type")
    r += 1

    make_checkbox_row(ws, r, 1, 8,
                      ["Ivoirienne", "Africaine", "Française", "Libanaise", "Asiatique", "Fast food", "Fusion"],
                      label="Cuisine")
    r += 1

    make_label_value(ws, r, 1, "Spécialité 1", 2, 4)
    make_label_value(ws, r, 5, "Spécialité 2", 6, 8)
    r += 1

    make_label_value(ws, r, 1, "Spécialité 3", 2, 4)
    make_label_value(ws, r, 5, "Plats sign.", 6, 8)
    r += 1

    make_checkbox_row(ws, r, 1, 8,
                      ["Carte photographiée", "Ardoise", "Verbal uniquement"],
                      label="Menu")
    r += 1

    # ═══════════════════════════════════════════════════
    # BLOC 3 — PRIX & RYTHME
    # ═══════════════════════════════════════════════════
    make_section_header(ws, r, 1, 8, "💰  PRIX & RYTHME", fill_blue)
    r += 1

    # Prix headers
    ws.cell(row=r, column=1).fill = fill_bg
    for ci, txt in [(2, ""), (3, "Min (F)"), (4, ""), (5, "Max (F)"), (6, ""), (7, ""), (8, "")]:
        c = ws.cell(row=r, column=ci, value=txt)
        c.font = Font(name="Calibri", bold=True, size=8, color=MUTED)
        c.alignment = center
        c.fill = fill_bg
    ws.row_dimensions[r].height = 14
    r += 1

    for label_txt in ["Éco / populaire", "Moyen / standard", "Premium / haut", "Boissons"]:
        make_label_value(ws, r, 1, label_txt, 2, 3)  # Min
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=4)
        ws.cell(row=r, column=4, value="→").font = Font(size=9, color=MUTED)
        ws.cell(row=r, column=4).alignment = center
        make_label_value(ws, r, 5, "", 5, 6)  # Max — no label needed since arrow
        # Patch: remove duplicate label, just use merged cells
        ws.cell(row=r, column=5).fill = fill_white
        ws.cell(row=r, column=5).border = border_thin
        ws.cell(row=r, column=6).border = border_thin
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        r += 1

    make_scale_row(ws, r, 1, "Budget", 2, 3, 8, "bas", "haut", 8)
    r += 1
    make_scale_row(ws, r, 1, "Cadence", 2, 3, 8, "lent", "rapide", 8)
    r += 1

    # ═══════════════════════════════════════════════════
    # BLOC 4 — LOGISTIQUE
    # ═══════════════════════════════════════════════════
    make_section_header(ws, r, 1, 8, "🏢  LOGISTIQUE", fill_grey)
    r += 1

    make_label_value(ws, r, 1, "Horaires L-V", 2, 3)
    make_label_value(ws, r, 4, "Samedi", 5, 6)
    make_label_value(ws, r, 7, "Dimanche", 8, 8)
    r += 1

    make_label_value(ws, r, 1, "Capacité int.", 2, 2)
    make_label_value(ws, r, 3, "Cap. ext.", 4, 4)
    make_label_value(ws, r, 5, "Parking", 6, 6, validation_key="parking")
    ws.row_dimensions[r].height = 22
    r += 1

    make_checkbox_row(ws, r, 1, 8,
                      ["Cash", "Orange Money", "MTN", "Moov", "Wave", "CB"],
                      label="Paiements")
    r += 1

    make_checkbox_row(ws, r, 1, 8,
                      ["Livraison", "À emporter", "Clim", "WiFi", "Terrasse", "Bar"],
                      label="Services")
    r += 1

    # ═══════════════════════════════════════════════════
    # BLOC 5 — DÉGUSTATION
    # ═══════════════════════════════════════════════════
    make_section_header(ws, r, 1, 8, "⭐  DÉGUSTATION", fill_red)
    r += 1

    # Testé ?
    ws.cell(row=r, column=1, value="Testé ?").font = label_font
    ws.cell(row=r, column=1).alignment = right_center
    ws.cell(row=r, column=1).fill = fill_bg
    ws.merge_cells(f"B{r}:C{r}")
    ws.cell(row=r, column=2, value="☐ Oui    ☐ Non").font = checkbox_font
    ws.cell(row=r, column=2).alignment = left_center
    ws.cell(row=r, column=2).fill = fill_white
    ws.cell(row=r, column=2).border = border_thin

    make_label_value(ws, r, 4, "Note /5", 5, 5)
    ws.merge_cells(f"F{r}:H{r}")
    ws.cell(row=r, column=6, value="☆  ☆  ☆  ☆  ☆").font = Font(size=16, color="D0D0D0")
    ws.cell(row=r, column=6).alignment = center
    ws.cell(row=r, column=6).fill = fill_white
    ws.cell(row=r, column=6).border = border_thin
    ws.row_dimensions[r].height = 26
    r += 1

    make_label_value(ws, r, 1, "Plats testés", 2, 8, row_height=28)
    r += 1

    make_label_value(ws, r, 1, "Budget test", 2, 3)
    ws.cell(row=r, column=4, value="FCFA").font = hint_font
    ws.cell(row=r, column=4).alignment = left_center
    r += 1

    make_label_value(ws, r, 1, "Points forts", 2, 8, row_height=28)
    r += 1
    make_label_value(ws, r, 1, "Points faibles", 2, 8, row_height=28)
    r += 1

    make_checkbox_row(ws, r, 1, 8,
                      ["🎯 Exact", "🤩 Surprise", "💸 Qualité-prix", "😐 Banal", "😞 Décevant"],
                      label="Ressenti")
    r += 1

    make_label_value(ws, r, 1, "Recommandation", 2, 3, validation_key="reco")
    make_label_value(ws, r, 4, "Accroche", 5, 8)
    r += 1

    # ═══════════════════════════════════════════════════
    # BLOC 6 — AMBIANCE & VERDICT
    # ═══════════════════════════════════════════════════
    make_section_header(ws, r, 1, 8, "🎭  AMBIANCE & VERDICT", fill_purple)
    r += 1

    make_scale_row(ws, r, 1, "Ambiance", 2, 3, 8, "simple", "raffiné", 8)
    r += 1
    make_scale_row(ws, r, 1, "Sociabilité", 2, 3, 8, "intime", "festif", 8)
    r += 1
    make_scale_row(ws, r, 1, "Notoriété", 2, 3, 8, "inconnu", "célèbre", 8)
    r += 1

    make_checkbox_row(ws, r, 1, 8,
                      ["Cadres/Pro", "Étudiants", "Familles", "Couples", "Solo", "Touristes", "Mixte"],
                      label="Clientèle")
    r += 1

    make_checkbox_row(ws, r, 1, 8,
                      ["Petit-déj", "Déjeuner", "Goûter", "Dîner", "Tard dans la nuit"],
                      label="Moments")
    r += 1

    make_label_value(ws, r, 1, "Affluence", 2, 3, validation_key="affluence")
    make_label_value(ws, r, 4, "Sonore", 5, 6, validation_key="sonore")
    r += 1

    make_label_value(ws, r, 1, "Propreté", 2, 3, validation_key="proprete")
    make_label_value(ws, r, 4, "Accueil", 5, 6, validation_key="accueil")
    r += 1

    # ── POTENTIEL — big box ──
    ws.merge_cells(f"A{r}:A{r+1}")
    c = ws.cell(row=r, column=1, value="POTENTIEL\nSPAWT")
    c.font = Font(name="Calibri", bold=True, size=10, color=GOLD)
    c.alignment = center
    c.fill = fill_light_gold
    c.border = border_box

    ws.merge_cells(f"B{r}:H{r}")
    pot_txt = "☐ 🔥 PÉPITE      ☐ ❤️ COUP DE CŒUR      ☐ ✅ SOLIDE      ☐ ⚠️ MOYEN      ☐ ❌ SKIP"
    c2 = ws.cell(row=r, column=2, value=pot_txt)
    c2.font = Font(name="Calibri", bold=True, size=10, color=DARK)
    c2.alignment = center
    c2.fill = fill_light_gold
    for ci in range(2, 9):
        ws.cell(row=r, column=ci).border = border_box
        ws.cell(row=r, column=ci).fill = fill_light_gold
    ws.row_dimensions[r].height = 28
    r += 1

    # Notes terrain
    ws.merge_cells(f"B{r}:H{r}")
    make_label_value(ws, r, 2, "", 2, 8, row_height=36)
    ws.cell(row=r, column=2).value = None
    ws.cell(row=r, column=2).font = value_font
    # mini label inside the merged POTENTIEL left cell is already done
    ws.row_dimensions[r].height = 36
    r += 1

    # ═══════════════════════════════════════════════════
    # BLOC 7 — B2B / COMMERCIAL
    # ═══════════════════════════════════════════════════
    make_section_header(ws, r, 1, 8, "🤝  B2B / COMMERCIAL", fill_grey)
    r += 1

    ws.cell(row=r, column=1, value="Contact ?").font = label_font
    ws.cell(row=r, column=1).alignment = right_center
    ws.cell(row=r, column=1).fill = fill_bg
    ws.merge_cells(f"B{r}:C{r}")
    ws.cell(row=r, column=2, value="☐ Oui    ☐ Non").font = checkbox_font
    ws.cell(row=r, column=2).alignment = left_center
    ws.cell(row=r, column=2).fill = fill_white
    ws.cell(row=r, column=2).border = border_thin
    make_label_value(ws, r, 4, "Nom", 5, 8)
    r += 1

    make_label_value(ws, r, 1, "Fonction", 2, 3, validation_key="fonction")
    make_label_value(ws, r, 4, "Réceptivité", 5, 6, validation_key="receptivite")
    r += 1

    ws.cell(row=r, column=1, value="RDV pris ?").font = label_font
    ws.cell(row=r, column=1).alignment = right_center
    ws.cell(row=r, column=1).fill = fill_bg
    ws.merge_cells(f"B{r}:C{r}")
    ws.cell(row=r, column=2, value="☐ Oui    ☐ Non").font = checkbox_font
    ws.cell(row=r, column=2).alignment = left_center
    ws.cell(row=r, column=2).fill = fill_white
    ws.cell(row=r, column=2).border = border_thin
    make_label_value(ws, r, 4, "Date RDV", 5, 6)

    make_checkbox_row(ws, r, 7, 8, ["Flyer", "Pitch", "CDV"])
    r += 1

    make_label_value(ws, r, 1, "Notes B2B", 2, 8, row_height=36)
    r += 1

    # ── Footer ──
    ws.merge_cells(f"A{r}:H{r}")
    c = ws.cell(row=r, column=1, value="SPAWT Collecte Terrain — Fiche à remplir sur place")
    c.font = Font(name="Calibri", size=8, color=MUTED, italic=True)
    c.alignment = center
    ws.row_dimensions[r].height = 18

    # ── Print area ──
    ws.print_area = f"A1:H{r}"


# ═══════════════════════════════════════════════════════════════
# Generate 20 fiches (20 sheets)
# ═══════════════════════════════════════════════════════════════
NB_FICHES = 20

for i in range(1, NB_FICHES + 1):
    if i == 1:
        ws = wb.active
        ws.title = f"Fiche {i:03d}"
    else:
        ws = wb.create_sheet(f"Fiche {i:03d}")
    ws.sheet_properties.tabColor = GOLD
    build_fiche(ws, i)

# ── Save ──
output = "/home/user/project_spawt/SPAWT_Fiches_Collecte.xlsx"
wb.save(output)
print(f"✅ {output}")
print(f"   → {NB_FICHES} fiches individuelles (1 onglet = 1 fiche)")
print(f"   → Format A4 portrait, prêt à imprimer")
print(f"   → Checkboxes à cocher, cases à remplir")
print(f"   → Listes déroulantes sur les champs clés")
